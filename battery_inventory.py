#!/usr/bin/env python3
"""
Battery Inventory — Titan AES
Run: python battery_inventory.py
Opens automatically at http://localhost:5555
"""
import base64, threading, webbrowser
from datetime import datetime
from io import BytesIO
from pathlib import Path

from flask import Flask, request, send_file, jsonify
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app  = Flask(__name__)
PORT = 5555

# ── Logo (loaded once at startup) ─────────────────────────────────────────────
_lp = Path(__file__).parent / 'Titanaes.png'
LOGO_URI = ("data:image/png;base64," + base64.b64encode(_lp.read_bytes()).decode()) if _lp.exists() else ""

# ── HTML ──────────────────────────────────────────────────────────────────────
HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>Battery Inventory · Titan AES</title>
<script src="https://unpkg.com/@zxing/browser@0.1.4/umd/index.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/qrcodejs/1.0.0/qrcode.min.js"></script>
<style>
:root{
  --primary:#0c2143; --primary-dark:#081630;
  --accent:#2563eb;  --accent-hover:#1d4ed8;
  --danger:#dc2626;  --danger-hover:#b91c1c;
  --success:#16a34a; --warn:#d97706;
  --bg:#f1f5f9; --card:#fff; --border:#e2e8f0;
  --text:#1e293b; --muted:#64748b;
}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;background:var(--bg);color:var(--text);min-height:100vh}

/* ── Header ── */
header{background:var(--primary);color:#fff;padding:10px 24px;display:flex;align-items:center;justify-content:space-between;box-shadow:0 2px 10px rgba(0,0,0,.25)}
.hdr-left{display:flex;align-items:center;gap:14px}
.hdr-left img{height:34px;display:block}
.hdr-div{width:1px;height:30px;background:rgba(255,255,255,.2)}
.hdr-title{font-size:1rem;font-weight:700;letter-spacing:.02em}
.hdr-date{font-size:.75rem;opacity:.6;margin-top:1px}
.hdr-right{display:flex;align-items:center;gap:8px}
.hdr-right label{font-size:.72rem;font-weight:600;opacity:.65;text-transform:uppercase;letter-spacing:.06em}
.hdr-right input{background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.25);color:#fff;border-radius:6px;padding:5px 10px;font-size:.85rem;outline:none;width:180px}
.hdr-right input::placeholder{opacity:.5}
.hdr-right input:focus{background:rgba(255,255,255,.2);border-color:rgba(255,255,255,.5)}

/* ── Stats Bar ── */
#stats-bar{display:none;background:#0f2d57;color:#fff;padding:8px 24px;gap:0;flex-wrap:wrap;font-size:.8rem}
.stat-seg{display:flex;align-items:center;gap:10px;padding:4px 20px;border-right:1px solid rgba(255,255,255,.15)}
.stat-seg:first-child{padding-left:0}
.stat-seg:last-child{border-right:none}
.stat-label{opacity:.6;font-weight:600;text-transform:uppercase;letter-spacing:.05em;font-size:.7rem}
.stat-val{font-weight:700;font-size:.88rem}
.stat-val span{opacity:.7;font-weight:400;font-size:.8rem;margin-left:2px}
.s-pass{color:#4ade80}.s-susp{color:#fbbf24}.s-fail{color:#f87171}

/* ── Layout ── */
.main{max-width:1340px;margin:0 auto;padding:20px 24px}
.panels{display:grid;grid-template-columns:370px 1fr;gap:20px;align-items:start}
@media(max-width:900px){.panels{grid-template-columns:1fr}}

/* ── Cards ── */
.card{background:var(--card);border-radius:12px;box-shadow:0 1px 4px rgba(0,0,0,.08);padding:20px;border:1px solid var(--border)}
.card-title{font-size:.92rem;font-weight:700;color:var(--primary);margin-bottom:16px;padding-bottom:10px;border-bottom:2px solid var(--border);display:flex;align-items:center;gap:8px}

/* ── Form ── */
.form-grid{display:grid;grid-template-columns:1fr 1fr;gap:12px}
.fg{display:flex;flex-direction:column;gap:4px}
.fg.full{grid-column:1/-1}
label{font-size:.7rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.06em}
.req{color:#e74c3c;margin-left:2px}
.input-wrap{position:relative;display:flex}
input[type=text],input[type=number],textarea,select{width:100%;padding:9px 12px;border:1.5px solid var(--border);border-radius:7px;font-size:.9rem;color:var(--text);background:#fff;transition:border-color .15s;outline:none}
input:focus,textarea:focus,select:focus{border-color:var(--accent);box-shadow:0 0 0 3px rgba(37,99,235,.1)}
textarea{resize:vertical;min-height:64px}
.scan-btn{position:absolute;right:5px;top:50%;transform:translateY(-50%);background:var(--accent);color:#fff;border:none;border-radius:5px;padding:5px 9px;font-size:.75rem;font-weight:600;cursor:pointer;white-space:nowrap;transition:background .15s}
.scan-btn:hover{background:var(--accent-hover)}
input.has-btn{padding-right:82px}
.help-tip{font-size:.7rem;color:var(--muted);margin-top:2px}
.actions{display:flex;gap:8px;margin-top:16px}

/* ── Buttons ── */
.btn{padding:9px 18px;border:none;border-radius:7px;font-size:.85rem;font-weight:600;cursor:pointer;transition:background .15s,transform .1s;display:inline-flex;align-items:center;gap:5px;white-space:nowrap}
.btn:active{transform:scale(.97)}
.btn-primary{background:var(--accent);color:#fff}.btn-primary:hover{background:var(--accent-hover)}
.btn-ghost{background:#e2e8f0;color:#475569}.btn-ghost:hover{background:#cbd5e1}
.btn-success{background:var(--success);color:#fff}.btn-success:hover{background:#15803d}
.btn-success:disabled{background:#94a3b8;cursor:not-allowed}
.btn-danger{background:var(--danger);color:#fff;padding:4px 10px;font-size:.76rem}.btn-danger:hover{background:var(--danger-hover)}
.btn-edit{background:#f1f5f9;color:#475569;border:1px solid var(--border);padding:4px 10px;font-size:.76rem}.btn-edit:hover{background:#e2e8f0}
.btn-icon{background:#f1f5f9;color:#475569;border:1px solid var(--border);padding:7px 12px;font-size:.82rem}.btn-icon:hover{background:#e2e8f0}
.btn-icon:disabled{opacity:.45;cursor:not-allowed}

/* ── Status ── */
.status{padding:8px 12px;border-radius:7px;font-size:.82rem;font-weight:500;display:none;margin-bottom:12px;align-items:center;gap:7px}
.status.info{background:#eff6ff;color:#1d4ed8;border:1px solid #bfdbfe}
.status.success{background:#f0fdf4;color:#16a34a;border:1px solid #bbf7d0}
.status.error{background:#fef2f2;color:#dc2626;border:1px solid #fecaca}

/* ── Toolbar ── */
.toolbar{display:flex;align-items:center;gap:8px;margin-bottom:12px;flex-wrap:wrap}
.toolbar-filter{flex:1;min-width:160px;position:relative}
.toolbar-filter input{padding:7px 10px 7px 32px;border:1.5px solid var(--border);border-radius:7px;font-size:.85rem;width:100%;outline:none}
.toolbar-filter input:focus{border-color:var(--accent)}
.toolbar-filter::before{content:'🔍';position:absolute;left:9px;top:50%;transform:translateY(-50%);font-size:.8rem;pointer-events:none}
.toolbar-btns{display:flex;gap:6px;flex-wrap:wrap}

/* ── Table ── */
.table-wrap{overflow-x:auto}
table{width:100%;border-collapse:collapse;font-size:.83rem}
thead tr{background:var(--primary);color:#fff}
thead th{padding:10px 12px;text-align:left;font-weight:600;white-space:nowrap;font-size:.78rem}
thead th.sortable{cursor:pointer;user-select:none}
thead th.sortable:hover{background:#1a3a6b}
thead th.sort-active{background:#1a3a6b}
tbody tr:nth-child(even){background:#f8fafc}
tbody tr:hover{background:#eff6ff}
tbody td{padding:8px 12px;border-bottom:1px solid var(--border);vertical-align:middle}
.badge{display:inline-block;background:var(--primary);color:#fff;padding:2px 9px;border-radius:99px;font-size:.76rem;font-weight:700}
.mono{font-family:'Courier New',monospace;font-size:.8rem}
.oos-cell{background:#fef3c7 !important;color:#92400e;font-weight:600}
.empty{text-align:center;color:var(--muted);padding:44px 20px;font-size:.88rem}
.empty-icon{font-size:2.2rem;margin-bottom:8px;display:block;opacity:.35}
.count-pill{background:var(--accent);color:#fff;border-radius:99px;min-width:20px;height:20px;padding:0 6px;font-size:.72rem;font-weight:700;display:inline-flex;align-items:center;justify-content:center}

/* ── Flag Badges ── */
.flag-pass{display:inline-block;background:#dcfce7;color:#15803d;padding:2px 9px;border-radius:99px;font-size:.76rem;font-weight:700;border:1px solid #bbf7d0}
.flag-suspect{display:inline-block;background:#fef3c7;color:#92400e;padding:2px 9px;border-radius:99px;font-size:.76rem;font-weight:700;border:1px solid #fde68a}
.flag-fail{display:inline-block;background:#fee2e2;color:#b91c1c;padding:2px 9px;border-radius:99px;font-size:.76rem;font-weight:700;border:1px solid #fecaca}

/* ── Inline Edit ── */
.tbl-input{padding:4px 7px;border:1.5px solid var(--accent);border-radius:5px;font-size:.82rem;color:var(--text);width:100%;outline:none;background:#eff6ff}
.tbl-input:focus{border-color:var(--accent-hover);box-shadow:0 0 0 2px rgba(37,99,235,.12)}

/* ── Toast ── */
#toast{position:fixed;bottom:24px;right:24px;background:#1e293b;color:#fff;padding:12px 16px;border-radius:10px;font-size:.85rem;display:none;align-items:center;gap:12px;box-shadow:0 4px 20px rgba(0,0,0,.3);z-index:2000;max-width:340px}
#toast-undo{background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.3);color:#fff;border-radius:6px;padding:4px 12px;font-size:.8rem;font-weight:600;cursor:pointer}
#toast-undo:hover{background:rgba(255,255,255,.25)}
#toast-close{background:none;border:none;color:rgba(255,255,255,.5);cursor:pointer;font-size:1rem;padding:0;margin-left:4px}

/* ── Modals ── */
.overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.7);z-index:1000;align-items:center;justify-content:center}
.overlay.open{display:flex}
.modal{background:#fff;border-radius:14px;padding:22px;width:460px;max-width:95vw;box-shadow:0 20px 60px rgba(0,0,0,.35)}
.modal h3{font-size:.98rem;font-weight:700;color:var(--primary);margin-bottom:14px}
.vid-wrap{position:relative;width:100%;aspect-ratio:4/3;background:#000;border-radius:9px;overflow:hidden;margin-bottom:10px}
#scan-video{width:100%;height:100%;object-fit:cover;display:block}
.scan-overlay{position:absolute;inset:0;display:flex;align-items:center;justify-content:center}
.scan-box{width:190px;height:140px;border:2.5px solid #60a5fa;border-radius:9px;box-shadow:0 0 0 9999px rgba(0,0,0,.38)}
.scan-line{position:absolute;width:186px;height:2px;background:linear-gradient(90deg,transparent,#60a5fa,transparent);animation:sweep 1.6s ease-in-out infinite}
@keyframes sweep{0%{top:calc(50% - 68px)}50%{top:calc(50% + 68px)}100%{top:calc(50% - 68px)}}
.scan-hint{text-align:center;font-size:.82rem;color:var(--muted);min-height:18px}
.modal-foot{display:flex;justify-content:flex-end;gap:8px;margin-top:14px}
.import-info{background:#f8fafc;border:1px solid var(--border);border-radius:8px;padding:14px;margin:12px 0;font-size:.88rem;line-height:1.6}
.import-info strong{color:var(--primary)}

/* ── Box styles ── */
.box-block{border:1.5px solid var(--border);border-radius:9px;padding:14px;margin-bottom:10px;background:#fafcff}
.box-block-hdr{display:flex;align-items:center;gap:8px;margin-bottom:10px}
.box-title{font-weight:700;font-size:.92rem;color:var(--primary);flex:1}
.box-count{background:#e2e8f0;color:#475569;padding:2px 9px;border-radius:99px;font-size:.72rem;font-weight:700}
.box-qr{display:flex;flex-direction:column;align-items:center;gap:6px;margin-top:8px}
.box-qr canvas{border:1px solid var(--border);border-radius:6px;padding:4px;background:#fff}
.box-qr-label{font-size:.7rem;color:var(--muted)}
.box-cells{margin-top:8px;font-size:.78rem;color:#475569;line-height:1.8;word-break:break-all}
.box-badge{display:inline-block;background:#dbeafe;color:#1d4ed8;padding:1px 7px;border-radius:99px;
  font-size:.74rem;font-weight:700;margin:1px}

/* ── Panel tabs ── */
.panel-tabs{display:flex;border-bottom:2px solid var(--border);margin:-20px -20px 16px;padding:0 20px}
.panel-tab{flex:1;padding:10px 4px;text-align:center;cursor:pointer;font-size:.8rem;font-weight:600;
  color:var(--muted);border-bottom:2.5px solid transparent;margin-bottom:-2px;transition:color .15s}
.panel-tab:hover{color:var(--text)}
.panel-tab.active{color:var(--accent);border-bottom-color:var(--accent)}

/* ── Defect types ── */
.defect-block{border:1.5px solid var(--border);border-radius:9px;padding:12px;margin-bottom:10px;background:#fafcff}
.defect-block-hdr{display:flex;align-items:center;gap:7px;margin-bottom:8px}
.defect-dot{width:11px;height:11px;border-radius:50%;flex-shrink:0}
.defect-name{font-weight:700;font-size:.85rem;color:var(--text);flex:1}
.defect-count{background:#e2e8f0;color:#475569;padding:2px 8px;border-radius:99px;font-size:.7rem;font-weight:700}
.defect-empty{text-align:center;color:var(--muted);padding:28px 10px;font-size:.82rem}
.defect-badge-cell{display:inline-block;padding:2px 9px;border-radius:99px;font-size:.74rem;font-weight:700;border:1px solid rgba(0,0,0,.08)}

/* ── Print ── */
.print-only{display:none}
@media print{
  .no-print{display:none!important}
  .print-only{display:block!important}
  header{background:var(--primary)!important;-webkit-print-color-adjust:exact;print-color-adjust:exact}
  #stats-bar{display:flex!important;-webkit-print-color-adjust:exact;print-color-adjust:exact}
  .main{padding:8px}
  .panels{grid-template-columns:1fr!important}
  .card:first-child{display:none!important}
  .card{box-shadow:none;border:1px solid #ccc;padding:12px}
  .card-title .count-pill,.card-title .btn,.toolbar{display:none!important}
  .card-title{border-bottom:1px solid #ccc}
  thead{background:var(--primary)!important;-webkit-print-color-adjust:exact;print-color-adjust:exact}
  tbody td:last-child,thead th:last-child{display:none}
  .flag-pass,.flag-suspect,.flag-fail{-webkit-print-color-adjust:exact;print-color-adjust:exact}
  .oos-cell{-webkit-print-color-adjust:exact;print-color-adjust:exact}
}
</style>
</head>
<body>

<header>
  <div class="hdr-left">
    <img src="__LOGO_URI__" alt="Titan AES"/>
    <div class="hdr-div"></div>
    <div>
      <div class="hdr-title">Battery Inventory</div>
      <div class="hdr-date" id="hdr-date"></div>
    </div>
  </div>
  <div class="hdr-right no-print">
    <label for="batch-name">Batch</label>
    <input id="batch-name" placeholder="e.g. 240116 Formation" oninput="saveState()" />
  </div>
</header>

<!-- Stats Bar -->
<div id="stats-bar">
  <div class="stat-seg">
    <div class="stat-label">Total</div>
    <div class="stat-val" id="stat-count">0</div>
  </div>
  <div class="stat-seg">
    <div class="stat-label">OCV</div>
    <div class="stat-val">
      avg <span id="stat-ocv-avg">—</span>V &nbsp;
      min <span id="stat-ocv-min">—</span> &nbsp;
      max <span id="stat-ocv-max">—</span>
    </div>
  </div>
  <div class="stat-seg">
    <div class="stat-label">Weight avg</div>
    <div class="stat-val"><span id="stat-wt-avg">—</span> g</div>
  </div>
  <div class="stat-seg">
    <div class="stat-label">Flag</div>
    <div class="stat-val">
      <span class="s-pass">✓ <span id="stat-pass">0</span></span> &nbsp;
      <span class="s-susp">⚠ <span id="stat-susp">0</span></span> &nbsp;
      <span class="s-fail">✗ <span id="stat-fail">0</span></span>
    </div>
  </div>
</div>

<div class="main">
  <div class="panels">

    <!-- ── Left: Form ── -->
    <div class="card no-print">
      <div class="panel-tabs">
        <div class="panel-tab active" id="ptab-add"    onclick="switchTab('add')">➕ Add</div>
        <div class="panel-tab"        id="ptab-defect" onclick="switchTab('defect')">🏷 Defects</div>
        <div class="panel-tab"        id="ptab-boxes"  onclick="switchTab('boxes')">📦 Boxes</div>
      </div>

      <!-- ── Tab: Add Battery ── -->
      <div id="tab-add">
      <div class="status" id="status"></div>
      <div class="form-grid">

        <div class="fg full">
          <label>Manufacturer ID <span class="req">*</span></label>
          <div class="input-wrap">
            <input type="text" id="mfgId" class="has-btn" placeholder="Scan or type…" autocomplete="off"/>
            <button class="scan-btn no-print" onclick="openScanner()">📷 Scan</button>
          </div>
          <div class="help-tip">USB scanner: click field then scan. Camera: click 📷</div>
        </div>

        <div class="fg">
          <label>Titan ID <span class="req">*</span></label>
          <input type="text" id="titanId" placeholder="e.g. 1"/>
          <div class="help-tip">Auto-increments. Edit to reorder.</div>
        </div>

        <div class="fg">
          <label>Flag</label>
          <select id="flag">
            <option value="Pass">✓ Pass</option>
            <option value="Suspect">⚠ Suspect</option>
            <option value="Fail">✗ Fail</option>
          </select>
        </div>

        <div class="fg">
          <label>OCV (V) <span class="req">*</span></label>
          <input type="number" id="ocv" step="0.001" min="0" max="10" placeholder="e.g. 3.312"/>
          <div class="help-tip">Kept between entries. ↑↓ = ±1 mV</div>
        </div>

        <div class="fg">
          <label>Weight (g) <span class="req">*</span></label>
          <input type="number" id="weight" step="0.1" min="0" placeholder="e.g. 1250.5"/>
          <div class="help-tip">Kept between entries.</div>
        </div>

        <div class="fg">
          <label>Box Number</label>
          <input type="text" id="boxNumber" placeholder="e.g. BOX-001"/>
          <div class="help-tip">Kept between entries.</div>
        </div>

        <div class="fg full">
          <label>Comments <span style="font-weight:400;text-transform:none;font-size:.7rem">(optional — Enter to add)</span></label>
          <textarea id="comments" placeholder="Leave blank to skip…" rows="2"></textarea>
        </div>

      </div>
      <div class="actions">
        <button class="btn btn-primary" onclick="addBattery()">Add Battery</button>
        <button class="btn btn-ghost"   onclick="clearForm()">Clear</button>
      </div>
      </div><!-- end tab-add -->

      <!-- ── Tab: Defect Types ── -->
      <div id="tab-defect" style="display:none">
        <div class="fg" style="margin-bottom:14px">
          <label>New Defect Type Name</label>
          <div style="display:flex;gap:6px">
            <input type="text" id="new-defect-name" placeholder="e.g. Dent, Low OCV, Leak…"
              onkeydown="if(event.key==='Enter')addDefectType()"/>
            <button class="btn btn-primary" onclick="addDefectType()" style="white-space:nowrap;padding:9px 14px">Add</button>
          </div>
        </div>
        <div id="defect-list">
          <div class="defect-empty" id="defect-empty-msg">
            <div style="font-size:1.8rem;margin-bottom:6px;opacity:.3">🏷</div>
            No defect types yet.<br>Add one above.
          </div>
        </div>
      </div><!-- end tab-defect -->

      <!-- ── Tab: Boxes ── -->
      <div id="tab-boxes" style="display:none">
        <div id="box-list">
          <div class="defect-empty" id="box-empty-msg">
            <div style="font-size:1.8rem;margin-bottom:6px;opacity:.3">📦</div>
            No boxes yet.<br>Add a Box Number when entering batteries.
          </div>
        </div>
      </div><!-- end tab-boxes -->

    </div>

    <!-- ── Right: List ── -->
    <div class="card">
      <div class="card-title">
        🔋 Battery List <span class="count-pill" id="count">0</span>
        <div style="flex:1"></div>
        <div class="toolbar-btns no-print" style="display:flex;gap:6px">
          <button class="btn btn-icon" onclick="document.getElementById('import-file').click()" title="Import Excel">⬆ Import</button>
          <button class="btn btn-icon" id="csv-btn"   onclick="exportCSV()"   disabled title="Export CSV">CSV</button>
          <button class="btn btn-success" id="export-btn" onclick="exportExcel()" disabled style="padding:7px 14px;font-size:.82rem">⬇ Excel</button>
          <button class="btn btn-icon" id="print-btn" onclick="window.print()" disabled title="Print">🖨</button>
        </div>
      </div>

      <!-- Filter toolbar -->
      <div class="toolbar no-print">
        <div class="toolbar-filter">
          <input type="text" id="filter-input" placeholder="Search Titan ID, Mfg ID, comments…" oninput="setFilter(this.value)"/>
        </div>
      </div>

      <div id="list-wrap">
        <div class="empty"><span class="empty-icon">🔍</span>No batteries recorded yet.<br>Scan or enter a Manufacturer ID to begin.</div>
      </div>
    </div>

  </div>
</div>

<!-- Scanner Modal -->
<div class="overlay no-print" id="scanner-overlay">
  <div class="modal">
    <h3>📷 Scan Barcode / QR Code</h3>
    <div class="vid-wrap">
      <video id="scan-video" autoplay muted playsinline></video>
      <div class="scan-overlay"><div class="scan-box"></div><div class="scan-line"></div></div>
    </div>
    <div class="scan-hint" id="scan-hint">Align barcode within the frame…</div>
    <div class="modal-foot"><button class="btn btn-ghost" onclick="closeScanner()">Cancel</button></div>
  </div>
</div>

<!-- Import Confirm Modal -->
<div class="overlay no-print" id="import-overlay">
  <div class="modal">
    <h3>⬆ Import Excel</h3>
    <div class="import-info">
      Found <strong id="import-count">0</strong> batteries in the file.<br>
      How would you like to add them to the list?
    </div>
    <div class="modal-foot">
      <button class="btn btn-ghost"    onclick="closeImport()">Cancel</button>
      <button class="btn btn-primary"  onclick="doImport('append')">Append to list</button>
      <button class="btn btn-danger"   onclick="doImport('replace')" style="padding:9px 14px">Replace list</button>
    </div>
  </div>
</div>

<!-- Toast -->
<div id="toast">
  <span id="toast-msg"></span>
  <button id="toast-undo">Undo</button>
  <button id="toast-close" onclick="hideToast()">✕</button>
</div>

<!-- Hidden file input for import -->
<input type="file" id="import-file" accept=".xlsx,.xls" style="display:none" onchange="handleImportFile(this)"/>

<script>
// ── State ──────────────────────────────────────────────────────────────────
let batteries    = [];
let nextId       = 1;
let sortCol      = null;
let sortDir      = 1;
let filterText   = '';
let undoBuffer   = null;
let undoTimer    = null;
let stTimer      = null;
let pendingImport= null;
let reader       = null;
let defectTypes  = [];

const DEFECT_COLORS = ['#3b82f6','#ef4444','#f59e0b','#10b981','#8b5cf6','#06b6d4','#f97316','#ec4899','#84cc16','#14b8a6'];

// ── Init ───────────────────────────────────────────────────────────────────
document.getElementById('hdr-date').textContent = new Date().toLocaleDateString('en-US',
  {weekday:'long',year:'numeric',month:'long',day:'numeric'});
loadState();
renderTable();
updateStats();
renderDefectList();

// ── Keyboard flow: Enter moves between form fields, final field triggers add ──
['mfgId','titanId','ocv','weight','boxNumber','comments'].forEach((id, i, arr) => {
  document.getElementById(id).addEventListener('keydown', e => {
    if (e.key !== 'Enter') return;
    e.preventDefault();
    if (i < arr.length - 1) document.getElementById(arr[i+1]).focus();
    else addBattery();
  });
});

// OCV ↑↓ = ±1 mV
document.getElementById('ocv').addEventListener('keydown', e => {
  if (e.key !== 'ArrowUp' && e.key !== 'ArrowDown') return;
  e.preventDefault();
  const el = e.target, dir = e.key === 'ArrowUp' ? 1 : -1;
  el.value = (Math.round(((parseFloat(el.value)||0)*1000) + dir) / 1000).toFixed(3);
});

// ── LocalStorage ───────────────────────────────────────────────────────────
function saveState() {
  try {
    localStorage.setItem('titan_inv', JSON.stringify({
      batteries, nextId, defectTypes,
      batchName: document.getElementById('batch-name').value || ''
    }));
  } catch(_) {}
}

function loadState() {
  try {
    const s = localStorage.getItem('titan_inv');
    if (!s) return;
    const d = JSON.parse(s);
    batteries   = (d.batteries   || []).map(b => ({flag:'Pass', defect:'', ...b}));
    defectTypes = (d.defectTypes || []);
    nextId      = d.nextId || 1;
    if (d.batchName) document.getElementById('batch-name').value = d.batchName;
    document.getElementById('titanId').value = nextId;
  } catch(_) {}
}

// ── Add Battery ────────────────────────────────────────────────────────────
function addBattery() {
  const mfgId     = document.getElementById('mfgId').value.trim();
  const titanId   = document.getElementById('titanId').value.trim();
  const ocv       = document.getElementById('ocv').value.trim();
  const weight    = document.getElementById('weight').value.trim();
  const flag      = document.getElementById('flag').value;
  const boxNumber = document.getElementById('boxNumber').value.trim();
  const comments  = document.getElementById('comments').value.trim();

  if (!mfgId)   { showStatus('Manufacturer ID is required.', 'error'); return; }
  if (!titanId) { showStatus('Titan ID is required.', 'error'); return; }
  if (!ocv)     { showStatus('OCV is required.', 'error'); return; }
  if (!weight)  { showStatus('Weight is required.', 'error'); return; }

  // Duplicate check
  const dup = batteries.find(b => b.mfgId === mfgId);
  if (dup) {
    showStatus(`Warning: Manufacturer ID already recorded as Titan #${dup.titanId}.`, 'error');
    // Still allow — just warn. User can confirm by clicking again.
    if (!window._dupConfirm) {
      window._dupConfirm = true;
      return;
    }
  }
  window._dupConfirm = false;

  batteries.push({
    titanId, mfgId, ocv: parseFloat(ocv), weight: parseFloat(weight),
    flag, boxNumber, defect:'', comments,
    date: new Date().toLocaleString('en-US',{year:'numeric',month:'2-digit',day:'2-digit',hour:'2-digit',minute:'2-digit',second:'2-digit'})
  });

  const n = parseInt(titanId);
  nextId = isNaN(n) ? nextId + 1 : n + 1;

  showStatus(`Battery "${titanId}" added. (${batteries.length} total)`, 'success');
  renderTable();
  clearForm();
}

function clearForm() {
  document.getElementById('mfgId').value    = '';
  document.getElementById('titanId').value  = nextId;
  document.getElementById('flag').value     = 'Pass';
  document.getElementById('comments').value = '';
  // OCV, Weight, Box intentionally kept
  document.getElementById('mfgId').focus();
}

// ── Delete / Undo ──────────────────────────────────────────────────────────
function deleteBattery(i) {
  undoBuffer = { item: {...batteries[i]}, idx: i };
  batteries.splice(i, 1);
  renderTable();
  showToast(`Battery #${undoBuffer.item.titanId} removed.`, undoDelete);
}

function undoDelete() {
  if (!undoBuffer) return;
  batteries.splice(undoBuffer.idx, 0, undoBuffer.item);
  undoBuffer = null;
  renderTable();
  showStatus('Deletion undone.', 'success');
}

// ── Edit Row ───────────────────────────────────────────────────────────────
function editRow(i) {
  const b = batteries[i];
  const row = document.getElementById('row-'+i);
  const flagOpts = ['Pass','Suspect','Fail'].map(f =>
    `<option value="${f}"${(b.flag||'Pass')===f?' selected':''}>${f}</option>`).join('');
  const defectOpts = ['', ...defectTypes.map(d=>d.name)].map(n =>
    `<option value="${n}"${(b.defect||'')===n?' selected':''}>${n||'— None —'}</option>`).join('');
  row.innerHTML = `
    <td><input class="tbl-input" id="e-tid"    value="${esc(b.titanId)}"    style="width:65px"/></td>
    <td><input class="tbl-input" id="e-mfg"    value="${esc(b.mfgId)}"     style="width:100%"/></td>
    <td><input class="tbl-input" id="e-ocv"    value="${b.ocv}"   type="number" step="0.001" style="width:72px"/></td>
    <td><input class="tbl-input" id="e-weight" value="${b.weight}" type="number" step="0.1"  style="width:72px"/></td>
    <td><select class="tbl-input" id="e-flag"   style="width:88px">${flagOpts}</select></td>
    <td><input class="tbl-input" id="e-box"    value="${esc(b.boxNumber||'')}"              style="width:80px"/></td>
    <td><select class="tbl-input" id="e-defect" style="width:100px">${defectOpts}</select></td>
    <td><input class="tbl-input" id="e-cmt"    value="${esc(b.comments||'')}"               style="width:100%"/></td>
    <td style="color:#64748b;font-size:.76rem;white-space:nowrap">${b.date}</td>
    <td style="white-space:nowrap">
      <button class="btn btn-primary" style="padding:4px 10px;font-size:.76rem" onclick="saveRow(${i})">Save</button>
      <button class="btn btn-ghost"   style="padding:4px 8px;font-size:.76rem"  onclick="renderTable()">✕</button>
    </td>`;
  document.getElementById('e-ocv').addEventListener('keydown', e => {
    if (e.key !== 'ArrowUp' && e.key !== 'ArrowDown') return;
    e.preventDefault();
    const el = e.target, dir = e.key === 'ArrowUp' ? 1 : -1;
    el.value = (Math.round(((parseFloat(el.value)||0)*1000) + dir) / 1000).toFixed(3);
  });
  document.getElementById('e-mfg').focus();
}

function saveRow(i) {
  const tid = document.getElementById('e-tid').value.trim();
  const mfg = document.getElementById('e-mfg').value.trim();
  const ocv = document.getElementById('e-ocv').value.trim();
  const wt  = document.getElementById('e-weight').value.trim();
  const flg = document.getElementById('e-flag').value;
  const box = document.getElementById('e-box').value.trim();
  const def = document.getElementById('e-defect').value;
  const cmt = document.getElementById('e-cmt').value.trim();
  if (!tid || !mfg || !ocv || !wt) { showStatus('All fields except Comments are required.', 'error'); return; }
  batteries[i] = {...batteries[i], titanId:tid, mfgId:mfg, ocv:parseFloat(ocv), weight:parseFloat(wt), flag:flg, boxNumber:box, defect:def, comments:cmt};
  renderTable();
  showStatus(`Battery "${tid}" updated.`, 'success');
}

// ── Stats ──────────────────────────────────────────────────────────────────
function updateStats() {
  const bar = document.getElementById('stats-bar');
  if (!batteries.length) { bar.style.display = 'none'; return; }
  bar.style.display = 'flex';

  const ocvs = batteries.map(b => parseFloat(b.ocv)).filter(v => !isNaN(v));
  const wts  = batteries.map(b => parseFloat(b.weight)).filter(v => !isNaN(v));
  const avg  = arr => arr.reduce((a,v) => a+v, 0) / arr.length;
  const flags = batteries.reduce((a,b) => { const f=b.flag||'Pass'; a[f]=(a[f]||0)+1; return a; }, {});

  document.getElementById('stat-count').textContent   = batteries.length;
  document.getElementById('stat-ocv-avg').textContent = ocvs.length ? avg(ocvs).toFixed(3) : '—';
  document.getElementById('stat-ocv-min').textContent = ocvs.length ? Math.min(...ocvs).toFixed(3) : '—';
  document.getElementById('stat-ocv-max').textContent = ocvs.length ? Math.max(...ocvs).toFixed(3) : '—';
  document.getElementById('stat-wt-avg').textContent  = wts.length  ? avg(wts).toFixed(1) : '—';
  document.getElementById('stat-pass').textContent    = flags['Pass']    || 0;
  document.getElementById('stat-susp').textContent    = flags['Suspect'] || 0;
  document.getElementById('stat-fail').textContent    = flags['Fail']    || 0;
}

// ── Sort / Filter ──────────────────────────────────────────────────────────
function toggleSort(col) {
  sortDir = (sortCol === col) ? -sortDir : 1;
  sortCol = col;
  renderTable();
}

function setFilter(text) {
  filterText = text;
  renderTable();
}

function getDisplayRows() {
  let rows = batteries.map((b, i) => ({...b, _idx: i}));
  if (filterText.trim()) {
    const q = filterText.trim().toLowerCase();
    rows = rows.filter(b =>
      String(b.titanId).toLowerCase().includes(q) ||
      (b.mfgId||'').toLowerCase().includes(q) ||
      (b.comments||'').toLowerCase().includes(q) ||
      (b.flag||'').toLowerCase().includes(q));
  }
  if (sortCol) {
    rows.sort((a, b) => {
      let av = a[sortCol], bv = b[sortCol];
      if (typeof av === 'string') { av = av.toLowerCase(); bv = String(bv).toLowerCase(); }
      return av < bv ? -sortDir : av > bv ? sortDir : 0;
    });
  }
  return rows;
}

function getOosIndices() {
  if (batteries.length < 4) return {ocv: new Set(), weight: new Set()};
  const ocvs = batteries.map(b => parseFloat(b.ocv));
  const wts  = batteries.map(b => parseFloat(b.weight));
  function ms(arr) {
    const m = arr.reduce((a,v)=>a+v,0)/arr.length;
    const s = Math.sqrt(arr.reduce((a,v)=>a+(v-m)**2,0)/arr.length);
    return {m, s};
  }
  const {m:om, s:os} = ms(ocvs);
  const {m:wm, s:ws} = ms(wts);
  const K = 2, ocvSet = new Set(), wtSet = new Set();
  batteries.forEach((_,i) => {
    if (os > 0 && Math.abs(ocvs[i]-om) > K*os) ocvSet.add(i);
    if (ws > 0 && Math.abs(wts[i]-wm)  > K*ws) wtSet.add(i);
  });
  return {ocv: ocvSet, weight: wtSet};
}

// ── Table ──────────────────────────────────────────────────────────────────
function renderTable() {
  updateStats();
  saveState();

  const n = batteries.length;
  document.getElementById('count').textContent     = n;
  document.getElementById('export-btn').disabled   = n === 0;
  document.getElementById('csv-btn').disabled      = n === 0;
  document.getElementById('print-btn').disabled    = n === 0;

  const wrap = document.getElementById('list-wrap');
  const rows = getDisplayRows();
  const oos  = getOosIndices();

  if (rows.length === 0) {
    wrap.innerHTML = n === 0
      ? `<div class="empty"><span class="empty-icon">🔍</span>No batteries recorded yet.<br>Scan or enter a Manufacturer ID to begin.</div>`
      : `<div class="empty"><span class="empty-icon">🔎</span>No results match "<strong>${esc(filterText)}</strong>".</div>`;
    return;
  }

  const COLS = [
    {key:'titanId',   label:'Titan ID'},
    {key:'mfgId',     label:'Manufacturer ID'},
    {key:'ocv',       label:'OCV (V)'},
    {key:'weight',    label:'Weight (g)'},
    {key:'flag',      label:'Flag'},
    {key:'boxNumber', label:'Box'},
    {key:'defect',    label:'Defect'},
    {key:'comments',  label:'Comments'},
    {key:'date',      label:'Date Added'},
  ];

  const hdrs = COLS.map(c => {
    const ico = sortCol===c.key ? (sortDir>0?' ▲':' ▼') : '';
    return `<th class="sortable${sortCol===c.key?' sort-active':''}" onclick="toggleSort('${c.key}')">${c.label}${ico}</th>`;
  }).join('') + '<th class="no-print"></th>';

  let tbody = '';
  rows.forEach(b => {
    const i = b._idx;
    const fc = {Pass:'flag-pass',Suspect:'flag-suspect',Fail:'flag-fail'}[b.flag||'Pass'] || 'flag-pass';
    const ocvCls = oos.ocv.has(i)    ? ' oos-cell' : '';
    const wtCls  = oos.weight.has(i) ? ' oos-cell' : '';
    const ocvTip = oos.ocv.has(i)    ? ' title="Out of spec (>2σ from batch avg)"' : '';
    const wtTip  = oos.weight.has(i) ? ' title="Out of spec (>2σ from batch avg)"' : '';
    const dt    = defectTypes.find(d => d.name === b.defect);
    const defectHtml = b.defect
      ? `<span class="defect-badge-cell" style="background:${dt?dt.color+'22':'#f1f5f9'};color:${dt?dt.color:'#475569'};border-color:${dt?dt.color+'55':'#e2e8f0'}">${esc(b.defect)}</span>`
      : '<span style="color:#94a3b8">—</span>';
    const boxHtml = b.boxNumber
      ? `<span class="box-badge">${esc(b.boxNumber)}</span>`
      : '<span style="color:#94a3b8">—</span>';
    tbody += `
      <tr id="row-${i}">
        <td><span class="badge">#${esc(b.titanId)}</span></td>
        <td><span class="mono">${esc(b.mfgId)}</span></td>
        <td class="${ocvCls}"${ocvTip}>${b.ocv}</td>
        <td class="${wtCls}"${wtTip}>${b.weight}</td>
        <td><span class="${fc}">${esc(b.flag||'Pass')}</span></td>
        <td>${boxHtml}</td>
        <td>${defectHtml}</td>
        <td>${b.comments ? esc(b.comments) : '<span style="color:#94a3b8">—</span>'}</td>
        <td style="white-space:nowrap;color:#64748b;font-size:.76rem">${b.date}</td>
        <td class="no-print" style="white-space:nowrap">
          <button class="btn btn-edit"   onclick="editRow(${i})">✏</button>
          <button class="btn btn-danger" onclick="deleteBattery(${i})">✕</button>
        </td>
      </tr>`;
  });

  wrap.innerHTML = `<div class="table-wrap"><table>
    <thead><tr>${hdrs}</tr></thead>
    <tbody>${tbody}</tbody>
  </table></div>`;
}

// ── Export Excel (server) ──────────────────────────────────────────────────
async function exportExcel() {
  if (!batteries.length) return;
  try {
    showStatus('Generating Excel…', 'info');
    const batch = document.getElementById('batch-name').value || '';
    const res = await fetch('/export', {
      method: 'POST',
      headers: {'Content-Type':'application/json'},
      body: JSON.stringify({batteries, batch})
    });
    if (!res.ok) throw new Error();
    const cd   = res.headers.get('Content-Disposition') || '';
    const fname= cd.match(/filename=(.+)/)?.[1] || 'Battery_Inventory.xlsx';
    const blob = await res.blob();
    const url  = URL.createObjectURL(blob);
    const a    = document.createElement('a');
    a.href = url; a.download = fname; a.click();
    URL.revokeObjectURL(url);
    showStatus(`Exported ${batteries.length} batteries to ${fname}`, 'success');
  } catch(_) { showStatus('Export failed.', 'error'); }
}

// ── Export CSV (client-side) ───────────────────────────────────────────────
function exportCSV() {
  if (!batteries.length) return;
  const hdr  = ['Titan ID','Manufacturer ID','OCV (V)','Weight (g)','Flag','Comments','Date Added'];
  const body = batteries.map(b =>
    [b.titanId, b.mfgId, b.ocv, b.weight, b.flag||'Pass', b.comments||'', b.date]);
  const csv = [hdr,...body].map(r =>
    r.map(v => `"${String(v).replace(/"/g,'""')}"`).join(',')).join('\r\n');
  const blob = new Blob(['\uFEFF'+csv], {type:'text/csv;charset=utf-8'});
  const url  = URL.createObjectURL(blob);
  const a    = document.createElement('a');
  a.href = url;
  a.download = `Battery_Inventory_${new Date().toISOString().slice(0,10)}.csv`;
  a.click();
  URL.revokeObjectURL(url);
}

// ── Import Excel ───────────────────────────────────────────────────────────
async function handleImportFile(input) {
  const file = input.files[0];
  if (!file) return;
  const fd = new FormData();
  fd.append('file', file);
  showStatus('Reading file…', 'info');
  try {
    const res  = await fetch('/import_excel', {method:'POST', body:fd});
    const data = await res.json();
    if (data.error) { showStatus('Import error: ' + data.error, 'error'); return; }
    if (!data.batteries.length) { showStatus('No battery data found in file.', 'error'); return; }
    pendingImport = data.batteries;
    document.getElementById('import-count').textContent = data.batteries.length;
    document.getElementById('import-overlay').classList.add('open');
  } catch(_) { showStatus('Import failed.', 'error'); }
  input.value = '';
}

function doImport(mode) {
  if (!pendingImport) return;
  if (mode === 'replace') {
    batteries = pendingImport.map(b => ({flag:'Pass', ...b}));
  } else {
    batteries.push(...pendingImport.map(b => ({flag:'Pass', ...b})));
  }
  nextId = batteries.reduce((m,b) => { const n=parseInt(b.titanId); return isNaN(n)?m:Math.max(m,n+1); }, nextId);
  document.getElementById('titanId').value = nextId;
  pendingImport = null;
  closeImport();
  renderTable();
  showStatus(`Imported. Total: ${batteries.length} batteries.`, 'success');
}

function closeImport() {
  document.getElementById('import-overlay').classList.remove('open');
}

// ── Scanner ────────────────────────────────────────────────────────────────
function openScanner() {
  document.getElementById('scanner-overlay').classList.add('open');
  document.getElementById('scan-hint').textContent = 'Starting camera…';
  startScan();
}
function closeScanner() { stopScan(); document.getElementById('scanner-overlay').classList.remove('open'); }

async function startScan() {
  if (typeof ZXingBrowser === 'undefined') {
    document.getElementById('scan-hint').textContent = 'Scanner unavailable. Use USB scanner or type manually.';
    return;
  }
  try {
    reader = new ZXingBrowser.BrowserMultiFormatReader();
    document.getElementById('scan-hint').textContent = 'Scanning — align barcode within the frame…';
    const result = await reader.decodeOnceFromVideoDevice(undefined, document.getElementById('scan-video'));
    document.getElementById('mfgId').value = result.getText();
    closeScanner();
    showStatus('Barcode scanned!', 'success');
    document.getElementById('titanId').focus();
  } catch(err) {
    if (err && err.name !== 'NotFoundException')
      document.getElementById('scan-hint').textContent = 'Camera unavailable. Type ID manually or use USB scanner.';
  }
}
function stopScan() { if (reader) { try { reader.reset(); } catch(_) {} reader = null; } }

document.getElementById('scanner-overlay').addEventListener('click', e => {
  if (e.target === document.getElementById('scanner-overlay')) closeScanner();
});
document.getElementById('import-overlay').addEventListener('click', e => {
  if (e.target === document.getElementById('import-overlay')) closeImport();
});

// ── Toast ──────────────────────────────────────────────────────────────────
function showToast(msg, undoFn) {
  clearTimeout(undoTimer);
  document.getElementById('toast-msg').textContent = msg;
  const undoBtn = document.getElementById('toast-undo');
  if (undoFn) { undoBtn.style.display='inline-flex'; undoBtn.onclick=()=>{undoFn();hideToast();}; }
  else          undoBtn.style.display='none';
  document.getElementById('toast').style.display = 'flex';
  undoTimer = setTimeout(hideToast, 5000);
}
function hideToast() {
  document.getElementById('toast').style.display = 'none';
  undoBuffer = null;
  clearTimeout(undoTimer);
}

// ── Status ─────────────────────────────────────────────────────────────────
function showStatus(msg, type) {
  const el = document.getElementById('status');
  el.className = 'status ' + type;
  el.textContent = msg;
  el.style.display = 'flex';
  clearTimeout(stTimer);
  stTimer = setTimeout(() => { el.style.display = 'none'; }, 4000);
}

// ── Boxes ──────────────────────────────────────────────────────────────────
function renderBoxes() {
  const el = document.getElementById('box-list');
  if (!el) return;

  // Group batteries by box
  const map = {};
  batteries.forEach(b => {
    const bn = (b.boxNumber || '').trim();
    if (!bn) return;
    if (!map[bn]) map[bn] = [];
    map[bn].push(b);
  });

  const boxes = Object.keys(map).sort();
  if (!boxes.length) {
    el.innerHTML = `<div class="defect-empty"><div style="font-size:1.8rem;margin-bottom:6px;opacity:.3">📦</div>No boxes yet.<br>Add a Box Number when entering batteries.</div>`;
    return;
  }

  el.innerHTML = boxes.map(bn => {
    const cells = map[bn];
    const cellBadges = cells.map(b => `<span class="box-badge">#${esc(b.titanId)}</span>`).join(' ');
    return `
      <div class="box-block">
        <div class="box-block-hdr">
          <span class="box-title">📦 ${esc(bn)}</span>
          <span class="box-count">${cells.length} cell${cells.length===1?'':'s'}</span>
          <button class="btn btn-ghost" style="padding:3px 9px;font-size:.72rem"
            onclick="filterByBox('${esc(bn)}')">View</button>
        </div>
        <div class="box-cells">${cellBadges}</div>
        <div class="box-qr">
          <div id="qr-${esc(bn).replace(/\s+/g,'-')}" style="line-height:0"></div>
          <span class="box-qr-label">${esc(bn)}</span>
          <button class="btn btn-ghost" style="padding:3px 9px;font-size:.72rem;margin-top:2px"
            onclick="printQR('${esc(bn)}')">🖨 Print label</button>
        </div>
      </div>`;
  }).join('');

  // Generate QR codes
  boxes.forEach(bn => {
    const divId = 'qr-' + bn.replace(/\s+/g,'-');
    const div   = document.getElementById(divId);
    if (div && typeof QRCode !== 'undefined') {
      div.innerHTML = '';
      new QRCode(div, { text: bn, width: 100, height: 100, correctLevel: QRCode.CorrectLevel.M });
    }
  });
}

function filterByBox(bn) {
  document.getElementById('filter-input').value = bn;
  setFilter(bn);
  switchTab('add');
  document.querySelector('[data-panel="list"]');
}

function printQR(bn) {
  const divId = 'qr-' + bn.replace(/\s+/g,'-');
  const div   = document.getElementById(divId);
  if (!div) return;
  const canvas = div.querySelector('canvas');
  if (!canvas) return;
  const win = window.open('', '_blank');
  win.document.write(`<html><body style="text-align:center;font-family:sans-serif;padding:30px">
    <img src="${canvas.toDataURL()}" style="width:180px;height:180px"><br>
    <p style="font-size:1.1rem;font-weight:700;margin-top:12px">${bn}</p>
    <script>window.onload=()=>{window.print();window.close()}<\/script>
  </body></html>`);
  win.document.close();
}

// ── Tab switching ─────────────────────────────────────────────────────────
function switchTab(tab) {
  document.getElementById('tab-add').style.display    = tab === 'add'    ? '' : 'none';
  document.getElementById('tab-defect').style.display = tab === 'defect' ? '' : 'none';
  document.getElementById('tab-boxes').style.display  = tab === 'boxes'  ? '' : 'none';
  document.getElementById('ptab-add').classList.toggle('active',    tab === 'add');
  document.getElementById('ptab-defect').classList.toggle('active', tab === 'defect');
  document.getElementById('ptab-boxes').classList.toggle('active',  tab === 'boxes');
  if (tab === 'boxes') renderBoxes();
}

// ── Defect types ───────────────────────────────────────────────────────────
function addDefectType() {
  const inp  = document.getElementById('new-defect-name');
  const name = inp.value.trim();
  if (!name) return;
  if (defectTypes.find(d => d.name.toLowerCase() === name.toLowerCase())) {
    showStatus('Defect type already exists.', 'error'); return;
  }
  defectTypes.push({
    id:    Date.now().toString(),
    name,
    color: DEFECT_COLORS[defectTypes.length % DEFECT_COLORS.length]
  });
  inp.value = '';
  renderDefectList();
  saveState();
}

function removeDefectType(id) {
  const dt = defectTypes.find(d => d.id === id);
  if (!dt) return;
  batteries.forEach(b => { if (b.defect === dt.name) b.defect = ''; });
  defectTypes = defectTypes.filter(d => d.id !== id);
  renderDefectList();
  renderTable();
  saveState();
}

function clearDefectAssignments(id) {
  const dt = defectTypes.find(d => d.id === id);
  if (!dt) return;
  batteries.forEach(b => { if (b.defect === dt.name) b.defect = ''; });
  renderDefectList();
  renderTable();
  saveState();
  showStatus(`Cleared all "${dt.name}" assignments.`, 'success');
}

function assignDefect(id) {
  const dt  = defectTypes.find(d => d.id === id);
  if (!dt) return;
  const raw = document.getElementById('dpaste-' + id).value;
  const ids = raw.split(/[\n,;\t]+/).map(s => s.trim()).filter(Boolean);
  if (!ids.length) { showStatus('Paste some IDs first.', 'error'); return; }
  let matched = 0;
  ids.forEach(qid => {
    const q = qid.toLowerCase();
    batteries.forEach(b => {
      if ((b.mfgId||'').toLowerCase() === q || String(b.titanId).toLowerCase() === q) {
        b.defect = dt.name; matched++;
      }
    });
  });
  document.getElementById('dpaste-' + id).value = '';
  renderDefectList();
  renderTable();
  saveState();
  showStatus(
    matched ? `Assigned "${dt.name}" to ${matched} batter${matched===1?'y':'ies'}.` : 'No matching IDs found.',
    matched ? 'success' : 'error'
  );
}

function renderDefectList() {
  const el    = document.getElementById('defect-list');
  const empty = document.getElementById('defect-empty-msg');
  if (!el) return;
  if (!defectTypes.length) {
    el.innerHTML = `<div class="defect-empty" id="defect-empty-msg"><div style="font-size:1.8rem;margin-bottom:6px;opacity:.3">🏷</div>No defect types yet.<br>Add one above.</div>`;
    return;
  }
  const counts = {};
  batteries.forEach(b => { if (b.defect) counts[b.defect] = (counts[b.defect]||0) + 1; });

  el.innerHTML = defectTypes.map(dt => `
    <div class="defect-block">
      <div class="defect-block-hdr">
        <div class="defect-dot" style="background:${dt.color}"></div>
        <span class="defect-name">${esc(dt.name)}</span>
        <span class="defect-count">${counts[dt.name]||0} cells</span>
        <button class="btn btn-ghost" style="padding:3px 8px;font-size:.7rem;margin-left:4px"
          onclick="clearDefectAssignments('${dt.id}')" title="Remove all assignments">Clear</button>
        <button class="btn btn-danger" style="padding:3px 8px;font-size:.7rem"
          onclick="removeDefectType('${dt.id}')" title="Delete defect type">✕</button>
      </div>
      <textarea id="dpaste-${dt.id}" rows="3"
        style="width:100%;font-size:.78rem;font-family:'Courier New',monospace;resize:vertical;
               padding:7px 9px;border:1.5px solid var(--border);border-radius:6px;outline:none"
        placeholder="Paste Manufacturer IDs here (one per line)…"></textarea>
      <button class="btn btn-primary" style="width:100%;margin-top:7px;padding:8px"
        onclick="assignDefect('${dt.id}')">Assign to Pasted IDs</button>
    </div>`).join('');
}

// ── Util ───────────────────────────────────────────────────────────────────
function esc(s) {
  return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}
</script>
</body>
</html>"""


# ── Excel Export ──────────────────────────────────────────────────────────────
@app.route('/export', methods=['POST'])
def export():
    data      = request.get_json(force=True)
    batteries = data.get('batteries', [])
    batch     = data.get('batch', '')
    now       = datetime.now()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Battery Inventory'

    C_NAVY, C_BLUE, C_STRIPE, C_WHITE = '0C2143', '2563EB', 'EFF6FF', 'FFFFFF'
    C_GREEN, C_AMBER, C_RED = 'DCFCE7', 'FEF3C7', 'FEE2E2'

    thin   = Side(style='thin', color='E2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(color): return PatternFill('solid', fgColor=color)
    def hdr(cell, bg=C_NAVY):
        cell.font      = Font(name='Calibri', bold=True, color=C_WHITE, size=11)
        cell.fill      = fill(bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Row 1: Title
    ws.merge_cells('A1:I1')
    ws['A1'] = 'Battery Inventory — Titan AES'
    ws['A1'].font      = Font(name='Calibri', size=14, bold=True, color=C_WHITE)
    ws['A1'].fill      = fill(C_NAVY)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Rows 2–4: Metadata
    meta = [('Export Date:', now.strftime('%Y-%m-%d %H:%M:%S')),
            ('Batch:',       batch or '—'),
            ('Total Count:', len(batteries))]
    for r, (k, v) in enumerate(meta, 2):
        ws.cell(r, 1, k).font = Font(bold=True, color='475569', size=10)
        ws.cell(r, 2, v).font = Font(color='1E293B', size=10)

    # Row 6: Column headers
    headers = ['Titan ID','Manufacturer ID','OCV (V)','Weight (g)','Flag','Box','Defect Type','Comments','Date Added']
    for c, h in enumerate(headers, 1):
        hdr(ws.cell(6, c, h), bg=C_BLUE)
    ws.row_dimensions[6].height = 20

    # Rows 7+: Data
    flag_fills = {'Pass': C_GREEN, 'Suspect': C_AMBER, 'Fail': C_RED}
    for ri, b in enumerate(batteries, 7):
        vals = [b['titanId'], b['mfgId'],
                float(b['ocv']), float(b['weight']),
                b.get('flag','Pass'), b.get('boxNumber',''),
                b.get('defect',''), b.get('comments',''), b['date']]
        stripe = fill(C_STRIPE if ri % 2 == 0 else C_WHITE)
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, v)
            cell.border    = border
            cell.alignment = Alignment(vertical='center')
            if ci == 5:  # Flag column
                cell.fill = fill(flag_fills.get(str(v), C_WHITE))
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 7 and v:  # Defect column
                cell.fill = fill('FEE2E2')
                cell.font = Font(bold=True, color='991B1B')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.fill = stripe

    # Header border
    for c in range(1, 10):
        ws.cell(6, c).border = border

    # Column widths
    for c, w in enumerate([12, 32, 11, 11, 12, 14, 18, 35, 22], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = 'A7'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Battery_Inventory_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Import Excel ──────────────────────────────────────────────────────────────
@app.route('/import_excel', methods=['POST'])
def import_excel():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file provided'}), 400
    try:
        wb   = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws   = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # Find header row
        hdr_idx = next((i for i, r in enumerate(rows)
                        if r and str(r[0]).strip() == 'Titan ID'), None)
        if hdr_idx is None:
            return jsonify({'error': 'Could not find "Titan ID" header row'}), 400

        hdrs = [str(h).strip() if h else '' for h in rows[hdr_idx]]
        col  = {h: i for i, h in enumerate(hdrs)}

        def gv(row, name, default=''):
            idx = col.get(name)
            return row[idx] if idx is not None and idx < len(row) else default

        result = []
        for row in rows[hdr_idx + 1:]:
            if not row or not row[0]:
                continue
            b = {
                'titanId':  str(gv(row,'Titan ID') or '').strip(),
                'mfgId':    str(gv(row,'Manufacturer ID') or '').strip(),
                'ocv':      float(gv(row,'OCV (V)', 0) or 0),
                'weight':   float(gv(row,'Weight (g)', 0) or 0),
                'flag':      str(gv(row,'Flag','Pass') or 'Pass').strip(),
                'boxNumber': str(gv(row,'Box','') or '').strip(),
                'defect':    str(gv(row,'Defect Type','') or '').strip(),
                'comments': str(gv(row,'Comments','') or '').strip(),
                'date':     str(gv(row,'Date Added','') or '').strip(),
            }
            if b['titanId']:
                result.append(b)

        return jsonify({'batteries': result, 'count': len(result)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Serve HTML ────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    html = HTML_TEMPLATE.replace('__LOGO_URI__', LOGO_URI)
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


# ── Launch ────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    import os
    host = os.environ.get('HOST', 'localhost')
    url  = f'http://localhost:{PORT}'
    print(f'  Battery Inventory  →  {url}')
    print('  Press Ctrl+C to stop.\n')
    if host == 'localhost':
        threading.Timer(1.2, lambda: webbrowser.open(url)).start()
    app.run(host=host, port=PORT, debug=False, use_reloader=False)
